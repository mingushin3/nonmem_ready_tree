"""Tier 0 auto-ingest adapter — happy/edge/trap + 974-regression guard.

합성 xlsx fixture는 test 시점에 openpyxl로 tmp_path에 생성(D-G3: 합성만, 실물 복사 금지,
바이너리 미커밋). 엔진/SSOT 무수정 — adapter는 run_strand를 무수정 호출하는 front-end.
assertion 스타일은 test_c_units.py(verbatim postcondition predicate) +
test_integration_slice9_batchB.py(run_strand record) 계승.
"""
import os

import openpyxl
import pytest

from src.adapter import ingest
from src.adapter.navigator import navigate, choose_detect_sequence
from src.adapter.column_canonicalizer import canonicalize
from src.adapter.xlsx_ingester import read_workbook_structure, build_engine_df
from src.orchestrator import REGISTRY, REQUIRES_DETECTION


# ----- 합성 fixture 빌더(tmp_path, 실물 mimic) ------------------------------

def _save(path, sheets):
    """sheets = [(title, [row, ...]), ...] → xlsx."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title=title)
        for r in rows:
            ws.append(r)
    wb.save(path)
    return str(path)


def _tidy_rows(extra_cols=None, conc_override=None):
    head = ["Subject", "Time", "Conc"] + (extra_cols or [])
    rows = [head]
    pad = [""] * len(extra_cols or [])
    data = [(1, 0, 0.0), (1, 1, 5.2), (2, 0, 0.0), (2, 1, 6.4)]
    for i, (sid, t, dv) in enumerate(data):
        c = conc_override[i] if conc_override else dv
        rows.append([sid, t, c] + pad)
    return rows


def make_tidy(tmp_path, **kw):
    return _save(tmp_path / "tidy.xlsx", [("data", _tidy_rows(**kw))])


# ===== HAPPY =================================================================

class TestHappyTidyLong:
    def test_full_fingerprint_band_completes(self, tmp_path):
        """clean tidy-long → faithful df, navigable 13c 완주, 구조 축 해소."""
        r = ingest(make_tidy(tmp_path))
        rec = r["dispatched"]["run_strand_record"]
        assert rec["boundary_at"] is None
        assert rec["terminal"] is None
        # 전 navigable 밴드가 dispatch됨(precondition 전부 충족)
        assert r["dispatched"]["c_sequence"] == [
            "c0201", "c0203", "c0205", "c0210", "c0211", "c0212",
            "c0214", "c0215", "c0216", "c0305", "c0310", "c0312", "c0314"]
        # 구조 파생 축 해소
        for ax in ("A1", "A3", "A5", "A10"):
            assert ax in r["axis_fingerprint"]["resolved"], ax
        # 정책 의존 축은 정직하게 undetermined(over-read 차단)
        assert "A0" in r["axis_fingerprint"]["undetermined"]
        assert r["entry_node"] == "N0"

    def test_c0212_postcondition_isinstance_bool(self, tmp_path):
        """c0212 postcondition predicate(verbatim): isinstance(meta.get('has_replicates'), bool)."""
        r = ingest(make_tidy(tmp_path))
        meta = r["dispatched"]["run_strand_record"]["meta"]
        assert isinstance(meta.get('has_replicates'), bool)

    def test_canonical_columns_present(self, tmp_path):
        """canonicalizer가 Subject/Time/Conc → subject_id/time_value/dv_value."""
        structure = read_workbook_structure(make_tidy(tmp_path))
        df_raw, _ = build_engine_df(structure["path"], structure)
        df, report = canonicalize(df_raw, structure)
        assert set(df.columns) >= {"subject_id", "time_value", "dv_value"}
        assert report["missing"] == []


# ===== EDGE =================================================================

class TestEdge:
    def test_blq_token_wired_reason_bound(self, tmp_path):
        """BLQ 토큰 tidy → c0305 wired reason(criterion_ko+ref) + blq_variants_found."""
        path = make_tidy(tmp_path, conc_override=[0.0, "BLQ", "<LLOQ", 6.4])
        r = ingest(path)
        wired = {x["feature"]: x for x in r["reasons"] if x["wired"]}
        assert "blq-token" in wired
        assert wired["blq-token"]["c_id"] == "c0305"
        assert wired["blq-token"]["criterion_ko"]  # 사람가독 사유 존재(M5)
        assert wired["blq-token"]["ref"]            # provenance 존재
        meta = r["dispatched"]["run_strand_record"]["meta"]
        assert isinstance(meta.get("blq_variants_found"), list)
        assert len(meta["blq_variants_found"]) >= 1  # anti-vacuous

    def test_cp949_korean_encoding_detected(self, tmp_path):
        """한글 텍스트 → c0216 has_encoding_issues=True + wired reason."""
        path = make_tidy(tmp_path, extra_cols=["비고"])
        # 비고 셀에 한글 주입
        wb = openpyxl.load_workbook(path)
        wb["data"]["D2"] = "재산출"
        wb.save(path)
        r = ingest(path)
        meta = r["dispatched"]["run_strand_record"]["meta"]
        assert isinstance(meta.get("has_encoding_issues"), bool)
        assert meta["has_encoding_issues"] is True
        assert any(x["feature"] == "encoding" and x["c_id"] == "c0216"
                   for x in r["reasons"] if x["wired"])

    def test_exact_duplicate_row_detected(self, tmp_path):
        """완전 중복 행 → c0215 has_exact_duplicates=True."""
        rows = _tidy_rows()
        rows.append(rows[1])  # happy 행 복제
        path = _save(tmp_path / "dup.xlsx", [("data", rows)])
        r = ingest(path)
        meta = r["dispatched"]["run_strand_record"]["meta"]
        assert isinstance(meta.get("has_exact_duplicates"), bool)
        assert meta["has_exact_duplicates"] is True

    def test_semistructured_multisheet_a10(self, tmp_path):
        """다중시트 → file_format=semi-structured 주입 → c0210 a10_state=SEMI-STRUCTURED."""
        path = _save(tmp_path / "multi.xlsx",
                     [("data", _tidy_rows()), ("notes", [["k", "v"], ["x", 1]])])
        r = ingest(path)
        meta = r["dispatched"]["run_strand_record"]["meta"]
        # c0210 postcondition predicate(verbatim 일부): a10_state ∈ 8 states
        assert meta.get("a10_state") in [
            'SDTM-ADaM', 'EDC-STRUCTURED', 'CRO-VENDOR', 'FLAT-TABULAR',
            'LEGACY-NM', 'SEMI-STRUCTURED', 'NON-TABULAR', 'CORRUPTED']
        assert meta["a10_state"] == "SEMI-STRUCTURED"

    def test_unit_separate_column_c0214_documented(self, tmp_path):
        """별도 Unit 열 → c0214 wired reason(derivable=False), df-default 정직 기록(R4/GAP-32)."""
        path = make_tidy(tmp_path, extra_cols=["Unit"])
        r = ingest(path)
        meta = r["dispatched"]["run_strand_record"]["meta"]
        assert meta.get("unit_declaration_complete") in (True, False)
        assert any(x["feature"] == "unit-column" and x["c_id"] == "c0214"
                   for x in r["reasons"] if x["wired"])


# ===== TRAP (honest-stop / annotate / two-option) ===========================

class TestTrapHonestStop:
    def test_subject_wide_honest_stop_a1_a10_only(self, tmp_path):
        """subject-as-column wide → df=None honest-stop, Fork-1: [c0201,c0210]만 dispatch."""
        rows = [["Time", "Animal1", "Animal2", "Animal3"],
                [0, 1.1, 2.2, 3.3], [1, 4.4, 5.5, 6.6]]
        path = _save(tmp_path / "wide.xlsx", [("conc", rows)])
        r = ingest(path)
        assert r["dispatched"]["c_sequence"] == ["c0201", "c0210"]
        # conc 의존 detector는 절대 dispatch 안 됨(precondition-gate, 날조 차단)
        for c in ("c0203", "c0205", "c0212", "c0305", "c0310"):
            assert c not in r["dispatched"]["c_sequence"], c
        assert r["stop"]["at"] == "structure-recognition"
        assert r["stop"]["gap"] == "GAP-37"
        assert any(x["feature"] == "subject-wide-conc" for x in r["reasons"] if not x["wired"])
        # 부분 fingerprint label로 over-read 차단
        assert "honest-stop" in r["axis_fingerprint"]["label"]
        assert set(r["axis_fingerprint"]["resolved"]) <= {"A1", "A10"}

    def test_param_summary_honest_stop(self, tmp_path):
        """파생 PK-param 요약(Parameters×Unit×Mean + 재산출) → honest-stop + GAP-37 주석."""
        rows = [
            ["", "", "G1", "G2", "G2 (재산출)"],
            ["", "", "7.5 mg/kg", "15 mg/kg", "15 mg/kg"],
            ["Parameters", "Unit", "Mean", "Mean", "Mean"],
            ["Cmax", "ng/mL", 23.4, 50.4, 50.4],
            ["AUClast", "ng*hr/mL", 665.6, 1339.0, 1339.0],
        ]
        path = _save(tmp_path / "param.xlsx", [("1. Data", rows)])
        r = ingest(path)
        assert r["stop"]["at"] == "structure-recognition"
        unwired = {x["feature"] for x in r["reasons"] if not x["wired"]}
        assert "param-summary" in unwired
        assert "mean-sd-aggregate" in unwired
        assert "reanalysis-duplicate" in unwired

    def test_intra_sheet_qa_block_honest_stop(self, tmp_path):
        """intra-sheet QA블록(Standard/DBLK/BLK) → honest-stop + GAP-37."""
        rows = [["Sample", "Conc"], ["Standard sample", 0], ["DBLK", 0],
                ["BLK (P)", 0], ["Subject-1", 5.2]]
        path = _save(tmp_path / "qa.xlsx", [("2. Result", rows)])
        r = ingest(path)
        assert r["stop"]["at"] == "structure-recognition"
        assert any(x["feature"] == "intra-sheet-qa-block" for x in r["reasons"] if not x["wired"])
        assert r["dispatched"]["c_sequence"] == ["c0201", "c0210"]

    def test_inline_blq_annotated_not_routed(self, tmp_path):
        """결합셀 '0.09 (BQL)' → unwired GAP-37 주석(wired c0305 토큰 regex가 놓치는 변종)."""
        path = make_tidy(tmp_path, conc_override=[0.0, "0.09 (BQL)", 5.2, 6.4])
        r = ingest(path)
        inline = [x for x in r["reasons"] if x["feature"] == "inline-blq"]
        assert inline and inline[0]["wired"] is False and inline[0]["gap"] == "GAP-37"

    def test_ambiguous_wide_two_interpretations(self, tmp_path):
        """재산출 마커 붙은 wide → 두 해석 surface(D-G2), 자동 선택 금지."""
        rows = [["Time", "G2", "G2 (재산출)"], [0, 1.1, 1.1], [1, 2.2, 2.3]]
        path = _save(tmp_path / "amb.xlsx", [("conc", rows)])
        r = ingest(path)
        dr = r["decision_required"]
        assert dr, "ambiguous-wide는 decision_required를 비우면 안 됨(silent 금지)"
        assert "interpretation_A" in dr[0] and "interpretation_B" in dr[0]
        # 모호 축에 대해 a-state를 날조하지 않음(resolved ⊆ file-property)
        assert set(r["axis_fingerprint"]["resolved"]) <= {"A1", "A10"}


# ===== REGRESSION GUARD (엔진/SSOT 불변) =====================================

class TestRegressionGuard:
    def test_adapter_import_does_not_mutate_registry(self, tmp_path):
        """ingest() 실행이 REGISTRY/REQUIRES_DETECTION을 변경하지 않음(adapter=front-end)."""
        before_reg = dict(REGISTRY)
        before_req = dict(REQUIRES_DETECTION)
        ingest(make_tidy(tmp_path))
        assert dict(REGISTRY) == before_reg
        assert dict(REQUIRES_DETECTION) == before_req

    def test_ingest_writes_no_spec_or_cunit_files(self, tmp_path):
        """ingest()가 spec/·src/c_units/에 파일을 쓰지 않음(read-only)."""
        root = os.path.dirname(os.path.dirname(__file__))
        snap = {d: sorted(os.listdir(os.path.join(root, d)))
                for d in ("spec", os.path.join("src", "c_units"))}
        ingest(make_tidy(tmp_path))
        after = {d: sorted(os.listdir(os.path.join(root, d)))
                 for d in ("spec", os.path.join("src", "c_units"))}
        assert after == snap

    def test_run_strand_contract_unmodified(self, tmp_path):
        """navigate가 엔진 run_strand를 무수정 소비 — record 키 == 정확한 계약."""
        structure = read_workbook_structure(make_tidy(tmp_path))
        df_raw, _ = build_engine_df(structure["path"], structure)
        df, _ = canonicalize(df_raw, structure)
        out = navigate(df, {"file_exists": True, "n_studies": 1}, faithful_tidy=True)
        assert set(out["run_strand_record"]) == {
            "actual_c_sequence", "total_cost", "terminal", "q_code",
            "boundary_at", "df", "meta"}

    def test_no_d_s1_runtime_error_on_detect_subsequence(self, tmp_path):
        """navigable 밴드는 전원 requires_detection_by=None → D-S1 RuntimeError 없음."""
        for c in ("c0201", "c0203", "c0205", "c0210", "c0211", "c0212",
                  "c0214", "c0215", "c0216", "c0305", "c0310", "c0312", "c0314"):
            assert REQUIRES_DETECTION.get(c) is None, c
        structure = read_workbook_structure(make_tidy(tmp_path))
        df_raw, _ = build_engine_df(structure["path"], structure)
        df, _ = canonicalize(df_raw, structure)
        out = navigate(df, {"file_exists": True, "n_studies": 1}, faithful_tidy=True)  # 무예외
        assert out["run_strand_record"]["boundary_at"] is None

    def test_honest_stop_choose_sequence_is_file_property_only(self):
        """choose_detect_sequence(df=None, faithful=False) == [c0201,c0210] (Fork-1 단위검증)."""
        seq = choose_detect_sequence(None, {"file_exists": True}, faithful_tidy=False)
        assert seq == ["c0210"]  # df=None이면 c0201(len(df)>0) 탈락, c0210(file_exists)만
        # 비어있지 않은 inventory-유사 df면 c0201도 포함
        import pandas as pd
        inv = pd.DataFrame({"sheet_name": ["a", "b"]})
        seq2 = choose_detect_sequence(inv, {"file_exists": True}, faithful_tidy=False)
        assert seq2 == ["c0201", "c0210"]


# ===== RECIPE-EMIT (경로 iii 보강: WU1 인벤토리 + 모델러 체크리스트, report-only) ========
# recipe = 기술(description)·안내일 뿐 — 변환·트리 라우팅·M2 무. 감지된 구조만 단정, 나머지는 verify.

def _wu(report, name):
    return [w for w in report["recipe"]["work_units"] if w["name"] == name]


class TestRecipeEmit:
    # ---- HAPPY -----------------------------------------------------------
    def test_recipe_status_described_not_executed(self, tmp_path):
        """clean tidy → recipe 존재, 최상위·모든 work_unit status == 'described, not executed'."""
        r = ingest(make_tidy(tmp_path))
        rec = r["recipe"]
        assert rec["status"] == "described, not executed"
        for w in rec["work_units"]:
            assert w["status"] == "described, not executed", w["name"]
        assert "M2 무의존" in rec["note"]  # 트리 라우팅·실행 안 함 명시

    def test_recipe_carries_commit_baseline(self, tmp_path):
        """추적성: recipe.commit_baseline에 코드 baseline 해시 2004d27 포함."""
        r = ingest(make_tidy(tmp_path))
        assert "2004d27" in r["recipe"]["commit_baseline"]

    # ---- EDGE (feature-driven WU) ---------------------------------------
    def test_qa_block_recipe_wu1_targets(self, tmp_path):
        """intra-sheet QA블록 → WU1 QA-strip, 제거대상에 Standard/DBLK/BLK 명시."""
        rows = [["Sample", "Conc"], ["Standard sample", 0], ["DBLK", 0],
                ["BLK (P)", 0], ["Subject-1", 5.2]]
        path = _save(tmp_path / "qa.xlsx", [("2. Result", rows)])
        r = ingest(path)
        wu1 = _wu(r, "QA-strip")
        assert wu1, "QA-strip WU1 필요"
        joined = " ".join(wu1[0]["targets"])
        assert "Standard" in joined and "DBLK" in joined and "BLK" in joined
        assert wu1[0]["sheets"] == ["2. Result"]
        assert wu1[0]["action"] == "remove"

    def test_param_summary_recipe_skip_reserve(self, tmp_path):
        """param-summary → WU1 skip-and-reserve(★ drop 아님), NCA 대조용 보존."""
        rows = [["", "", "G1", "G2"], ["Parameters", "Unit", "Mean", "Mean"],
                ["Cmax", "ng/mL", 23.4, 50.4]]
        path = _save(tmp_path / "param.xlsx", [("1. Data", rows)])
        r = ingest(path)
        wu = _wu(r, "param-summary-reserve")
        assert wu, "param-summary-reserve WU 필요"
        assert wu[0]["action"] == "skip-and-reserve"
        assert "drop" not in wu[0]["action"]
        assert "NCA" in wu[0]["note"]

    def test_bw_sheet_recipe_wu4_join(self, tmp_path):
        """BW 시트 + conc → WU4 join, bw_source.in_workbook=True, sheets=[BW]."""
        path = _save(tmp_path / "bw.xlsx", [
            ("2. Result", [["Sample", "Conc"], ["Standard sample", 0], ["Subject-1", 5.2]]),
            ("BW", [["Animal", "BW"], ["A1", 10.2]])])
        r = ingest(path)
        wu4 = _wu(r, "dose-bw-join")
        assert wu4
        assert wu4[0]["bw_source"]["in_workbook"] is True
        assert wu4[0]["bw_source"]["sheets"] == ["BW"]

    def test_no_bw_sheet_recipe_external(self, tmp_path):
        """conc 있으나 BW 시트 부재 → WU4 external(in_workbook=False), 외부 PDF는 '확인' 안내."""
        rows = [["Sample", "Conc"], ["Standard sample", 0], ["Subject-1", 5.2]]
        path = _save(tmp_path / "qa.xlsx", [("2. Result", rows)])
        r = ingest(path)
        wu4 = _wu(r, "dose-bw-join")
        assert wu4
        assert wu4[0]["bw_source"]["in_workbook"] is False
        assert any("PDF" in v for v in wu4[0]["verify"])  # 외부 출처는 예시·확인 안내(존재 단정 아님)

    def test_modeler_checklist_flag_only(self, tmp_path):
        """BLQ 토큰 → checklist blq-zero-policy decided=False; comparator 미점화(RLD 마커 없음)."""
        path = make_tidy(tmp_path, conc_override=[0.0, "BLQ", "<LLOQ", 6.4])
        r = ingest(path)
        ids = {it["id"]: it for it in r["recipe"]["checklist"]}
        assert "blq-zero-policy" in ids
        assert ids["blq-zero-policy"]["decided"] is False     # adapter 자동결정 절대 금지
        assert "comparator-arm-exclusion" not in ids          # RLD/Advagraf 마커 없음 → 미점화(정직)

    def test_comparator_marker_flags_advagraf(self, tmp_path):
        """파일명 'RLD 비교' → checklist comparator-arm-exclusion + evidence(자동결정 아님)."""
        rows = [["Sample", "Conc"], ["Standard sample", 0], ["Subject-1", 5.2]]
        path = _save(tmp_path / "beagle (RLD 비교).xlsx", [("2. Result", rows)])
        r = ingest(path)
        comp = [it for it in r["recipe"]["checklist"] if it["id"] == "comparator-arm-exclusion"]
        assert comp, "RLD 마커 → comparator 체크리스트 필요"
        assert comp[0]["decided"] is False
        assert comp[0]["evidence"]  # 근거 동반(cite-verify)

    # ---- TRAP (검토 §3 비결정 교정) -------------------------------------
    def test_param_summary_reanalysis_is_non_decision(self, tmp_path):
        """★검토 §3: param-summary 재산출 → 비결정(파생 재적합), arm-vs-replicate 2옵션 아님."""
        rows = [["", "", "G2", "G2 (재산출)"], ["Parameters", "Unit", "Mean", "Mean"],
                ["Cmax", "ng/mL", 50.4, 50.4]]
        path = _save(tmp_path / "param.xlsx", [("1. Data", rows)])
        r = ingest(path)
        nd = r["non_decisions"]
        assert nd and nd[0]["kind"] == "derived-parameter-refit"
        assert nd[0]["classification"] == "non-decision"
        assert nd[0]["evidence"]  # 재산출 셀 근거(cite-verify)
        # arm-vs-replicate 2옵션이 param-summary 시트에 대해 surface되지 않음(오분류 제거)
        assert not any(d.get("sheet") == "1. Data" for d in r["decision_required"])

    def test_two_option_preserved_for_nonparam_wide(self, tmp_path):
        """가드 과적용 방지: param-summary 아닌 wide(재산출)는 2옵션 그대로 유지(D-G2)."""
        rows = [["Time", "G2", "G2 (재산출)"], [0, 1.1, 1.1], [1, 2.2, 2.3]]
        path = _save(tmp_path / "amb.xlsx", [("conc", rows)])
        r = ingest(path)
        dr = r["decision_required"]
        assert dr and "interpretation_A" in dr[0] and "interpretation_B" in dr[0]
        assert r["non_decisions"] == []  # 비-param wide는 비결정으로 빼돌리지 않음

    # ---- HONESTY (단정 vs 확인) -----------------------------------------
    def test_recipe_verify_not_assert(self, tmp_path):
        """QA에 가려진 conc → WU3 wide layout 단정 안 함, conc 출처는 verify(='확인하라')."""
        rows = [["Sample", "Conc"], ["Standard sample", 0], ["Subject-1", 5.2]]
        path = _save(tmp_path / "qa.xlsx", [("2. Result", rows)])
        r = ingest(path)
        wu3 = _wu(r, "pivot-wide-to-long")
        assert wu3
        assert wu3[0]["wide_layout_detected"] is False           # 단정 금지(자동확정 못 함)
        assert any("출처" in v and "확인" in v for v in wu3[0]["verify"])  # conc 출처는 '확인하라'

    def test_recipe_emit_keeps_honest_stop(self, tmp_path):
        """recipe-emit는 변환 안 함 — honest-stop 유지(file-property만, df 미환원)."""
        rows = [["Sample", "Conc"], ["Standard sample", 0], ["Subject-1", 5.2]]
        path = _save(tmp_path / "qa.xlsx", [("2. Result", rows)])
        r = ingest(path)
        assert r["dispatched"]["c_sequence"] == ["c0201", "c0210"]  # recipe 생겨도 추측 dispatch 0
        assert r["stop"]["at"] == "structure-recognition"
