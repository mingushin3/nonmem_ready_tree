"""xlsx_ingester — 실물 xlsx → (구조 사실, 충실 df 또는 None). Tier 0 1단계.

openpyxl로 시트/병합/헤더 영역을 구조적으로 판독하고, 어느 시트가 tidy-long
(행=레코드, 1행=헤더)으로 **충실히** 환원되는지 판정한다. 충실 환원 불가
(subject-wide·param-summary·QA-contaminated·unknown)면 df=None을 반환한다.
  ★ pivot/melt/header-skip로 조립하지 않는다 — 그건 미배선 front-half transform
    (c0120 PIVOT 등, Tier 0 밖)이며 silent 조립은 날조다(D-G2).

엔진/SSOT 무수정. 합성·실물 모두 read-only. cp949 우려는 .xlsx에 무해(OOXML 내부 UTF-8);
비-ASCII 존재는 c0216 ENCODING 신호로만 기록한다(decode 시도 안 함).
"""
import re
from pathlib import Path

import openpyxl
import pandas as pd

_SAMPLE_ROWS = 20
_SAMPLE_COLS = 12

# intra-sheet QA 블록 토큰(실샘플 위 동일시트 혼재 = not-accepted)
_QA_TOKENS = ("standard", "dblk", " blk", "blk ", "lqc", "mqc", "hqc",
              "calibration", " qc", "qc ", "blank", "dilution", "spike")
# 시트 분류용 헤더 패턴(canonicalizer와 의도적으로 독립 — 각 c가 자체 컬럼상수를 갖는 house style)
_TIME_HDR = re.compile(r"(time|tad|tafd|시간|nominal)", re.I)
_DV_HDR = re.compile(r"(conc|concentration|\bdv\b|농도|observation)", re.I)
_SUBJ_HDR = re.compile(r"(subject|animal|개체|동물|^id$|^#\s*\d+$|^m\d+$)", re.I)


def is_lock_file(path) -> bool:
    """엑셀 잠금파일(~$...xlsx)은 워크북이 아님 — openpyxl이 에러내므로 사전 필터."""
    return Path(path).name.startswith("~$")


def _s(v) -> str:
    return "" if v is None else str(v)


def _has_non_ascii(values) -> bool:
    return any(any(ord(ch) > 127 for ch in _s(v)) for v in values)


def _looks_qa(sample_rows) -> bool:
    for r in sample_rows:
        if r and r[0] is not None and any(t in _s(r[0]).lower() for t in _QA_TOKENS):
            return True
    return False


def _looks_param_summary(sample_rows) -> bool:
    for r in sample_rows:
        cells = [_s(v).strip().lower() for v in r if v is not None]
        if any("parameter" in c for c in cells) and any(c in ("unit", "units") for c in cells):
            return True
    return False


def _header_cells(header_row):
    return [_s(v).strip() for v in header_row if v is not None and _s(v).strip()]


def _looks_subject_wide(header_row) -> bool:
    # 1행 헤더에 subject/animal류 컬럼이 2개 이상 = subject-as-column wide
    return sum(1 for h in _header_cells(header_row) if _SUBJ_HDR.search(h)) >= 2


def _looks_tidy(header_row) -> bool:
    hdr = _header_cells(header_row)
    return any(_TIME_HDR.search(h) for h in hdr) and any(_DV_HDR.search(h) for h in hdr)


def _classify_shape(header_row, sample_rows) -> str:
    """우선순위: qa > param-summary > subject-wide > tidy-long > unknown.

    (2.Result는 QA와 deeper-subject-wide를 둘 다 가지므로 QA 우선 = qa-contaminated.
     1행만 헤더로 본다 — 헤더가 metadata rows 아래 묻혀 있으면 tidy 아님(정직, header-skip은 front-half).)
    """
    if _looks_qa(sample_rows):
        return "qa-contaminated"
    if _looks_param_summary(sample_rows):
        return "param-summary"
    if _looks_subject_wide(header_row):
        return "subject-wide"
    if _looks_tidy(header_row):
        return "tidy-long"
    return "unknown"


def read_workbook_structure(xlsx_path: str) -> dict:
    """xlsx → 시트별 구조 사실(dims·merged·shape_class·top-left sample)."""
    if is_lock_file(xlsx_path):
        raise ValueError(f"lock file, not a workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    per_sheet = {}
    non_ascii = _has_non_ascii(wb.sheetnames)
    for ws in wb.worksheets:
        sample = list(ws.iter_rows(min_row=1, max_row=_SAMPLE_ROWS,
                                   max_col=_SAMPLE_COLS, values_only=True))
        header_row = sample[0] if sample else ()
        if _has_non_ascii([v for row in sample for v in row]):
            non_ascii = True
        per_sheet[ws.title] = {
            "n_rows": ws.max_row,
            "n_cols": ws.max_column,
            "n_merged": len(ws.merged_cells.ranges),
            "shape_class": _classify_shape(header_row, sample),
            "header_row": [_s(v) for v in header_row],
            "sample": [[_s(v) for v in row] for row in sample],
        }
    names = list(wb.sheetnames)
    wb.close()
    return {
        "path": str(xlsx_path),
        "sheet_names": names,
        "n_sheets": len(names),
        "per_sheet": per_sheet,
        "non_ascii_present": bool(non_ascii),
        "encoding_hint": "cp949-suspected" if non_ascii else "ascii",
    }


def _sheet_to_df(xlsx_path: str, sheet_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    header, seen = [], {}
    for i, h in enumerate(rows[0]):
        name = _s(h).strip() or f"col{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}.{seen[name]}"
        else:
            seen[name] = 0
        header.append(name)
    df = pd.DataFrame(rows[1:], columns=header)
    return df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)


def build_engine_df(xlsx_path: str, structure: dict):
    """tidy-long 시트가 있으면 (df, build_report:faithful=True). 없으면 (None, faithful=False).

    ★ 절대 pivot/melt 안 함 — subject-wide·param-summary는 df=None으로 정직 정지시킨다.
    """
    shapes = {n: s["shape_class"] for n, s in structure["per_sheet"].items()}
    tidy = [n for n, sc in shapes.items() if sc == "tidy-long"]
    if not tidy:
        return None, {
            "chosen_sheet": None,
            "faithful": False,
            "reason": "tidy-long으로 충실 환원되는 시트 없음 (구조 조립 front-half 미배선)",
            "sheet_shapes": shapes,
            "raw_columns": [],
        }
    chosen = tidy[0]
    df = _sheet_to_df(xlsx_path, chosen)
    return df, {
        "chosen_sheet": chosen,
        "faithful": True,
        "reason": None,
        "sheet_shapes": shapes,
        "raw_columns": list(df.columns),
    }
